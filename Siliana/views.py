from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Produit, Vente, Achat, Order, OrderItem
from datetime import date
from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def user_login(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
    
    return render(request, 'login.html')


def user_logout(request):
    logout(request)
    messages.success(request, 'تم تسجيل الخروج بنجاح')
    return redirect('login')


@login_required
def home(request):
    total_products = Produit.objects.count()
    sales_today = Vente.objects.filter(date_vente=date.today()).count()
    
    ventes_today = Vente.objects.filter(date_vente=date.today())
    revenue_today = sum(vente.total() for vente in ventes_today)
    
    low_stock_products = Produit.objects.filter(quantite__lte=5)
    low_stock = low_stock_products.count()
    
    context = {
        'total_products': total_products,
        'sales_today': sales_today,
        'revenue_today': revenue_today,
        'low_stock': low_stock,
        'low_stock_products': low_stock_products[:5]
    }
    return render(request, 'home.html', context)


@login_required
def product_list(request):
    products = Produit.objects.all().order_by('nom')
    return render(request, 'product_list.html', {'products': products})


@login_required
def add_product(request):
    if request.method == 'POST':
        nom = request.POST.get('nom')
        quantite = request.POST.get('quantite', 0)
        prix_achat = request.POST.get('prix_achat')
        prix_vente = request.POST.get('prix_vente')
        
        Produit.objects.create(
            nom=nom,
            quantite=quantite,
            prix_achat=prix_achat,
            prix_vente=prix_vente
        )
        messages.success(request, f'تم إضافة المنتج "{nom}" بنجاح')
        return redirect('product_list')
    
    return render(request, 'product_form.html')


@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Produit, id=product_id)
    
    if request.method == 'POST':
        product.nom = request.POST.get('nom')
        product.quantite = request.POST.get('quantite')
        product.prix_achat = request.POST.get('prix_achat')
        product.prix_vente = request.POST.get('prix_vente')
        product.save()
        
        messages.success(request, f'تم تحديث المنتج "{product.nom}" بنجاح')
        return redirect('product_list')
    
    return render(request, 'product_form.html', {'product': product})


@login_required
def add_stock(request, product_id):
    product = get_object_or_404(Produit, id=product_id)
    
    if request.method == 'POST':
        quantite = int(request.POST.get('quantite', 0))
        
        if quantite > 0:
            Achat.objects.create(produit=product, quantite=quantite)
            messages.success(request, f'تم إضافة {quantite} قطعة من "{product.nom}" إلى المخزون')
            return redirect('product_list')
        else:
            messages.error(request, 'يجب إدخال كمية صحيحة')
    
    return render(request, 'add_stock.html', {'product': product})


@login_required
def new_sale(request):
    products = Produit.objects.filter(quantite__gt=0).order_by('nom')
    
    if request.method == 'POST':
        produit_id = request.POST.get('produit')
        quantite = int(request.POST.get('quantite', 0))
        
        try:
            produit = get_object_or_404(Produit, id=produit_id)
            
            if quantite <= 0:
                messages.error(request, 'يجب إدخال كمية صحيحة')
            elif quantite > produit.quantite:
                messages.error(request, f'الكمية المتوفرة فقط {produit.quantite} قطعة')
            else:
                vente = Vente.objects.create(produit=produit, quantite=quantite)
                total = vente.total()
                messages.success(request, f'تمت عملية البيع بنجاح! الإجمالي: {total} د.ت')
                return redirect('new_sale')
        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    
    return render(request, 'new_sale.html', {'products': products})


@login_required
def sales_report(request):
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    ventes = Vente.objects.all().order_by('-date_vente')
    
    if date_debut and date_fin:
        ventes = ventes.filter(date_vente__range=[date_debut, date_fin])
    
    total = sum(vente.total() for vente in ventes)
    
    return render(request, 'sales_report.html', {
        'ventes': ventes,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'total': total
    })


def public_order(request):
    """Public order page for unlogged users"""
    produits = Produit.objects.all().order_by('nom')
    
    if request.method == 'POST':
        nom = request.POST.get('nom')
        wilaya = request.POST.get('wilaya')
        ville = request.POST.get('ville')
        telephone = request.POST.get('telephone')
        
        # Validate required fields
        if not all([nom, wilaya, ville, telephone]):
            messages.error(request, 'الرجاء ملء جميع الحقول المطلوبة')
            return render(request, 'public_order.html', {'produits': produits})
        
        # Create order with transaction
        try:
            with transaction.atomic():
                order = Order.objects.create(
                    nom=nom,
                    wilaya=wilaya,
                    ville=ville,
                    telephone=telephone
                )
                
                has_items = False
                for produit in produits:
                    quantity_key = f'product_{produit.id}'
                    quantite = request.POST.get(quantity_key, '0')
                    
                    try:
                        quantite = int(quantite)
                    except ValueError:
                        quantite = 0
                    
                    if quantite > 0:
                        if quantite <= produit.quantite:
                            OrderItem.objects.create(
                                order=order,
                                produit=produit,
                                quantite=quantite,
                                prix=produit.prix_vente
                            )
                            has_items = True
                        else:
                            messages.error(request, f'الكمية المطلوبة من {produit.nom} غير متوفرة')
                            return render(request, 'public_order.html', {'produits': produits})
                
                if not has_items:
                    order.delete()
                    messages.error(request, 'الرجاء اختيار منتج واحد على الأقل')
                    return render(request, 'public_order.html', {'produits': produits})
                
                messages.success(request, '✓ تم إرسال طلبك بنجاح! سنتصل بك قريباً')
                return redirect('public_order')
                
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    
    return render(request, 'public_order.html', {'produits': produits})


@login_required
def orders_list(request):
    """Orders list page for logged users only"""
    status_filter = request.GET.get('status', '')
    wilaya_filter = request.GET.get('wilaya', '')
    
    orders = Order.objects.all().order_by('-date_commande')
    
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if wilaya_filter:
        orders = orders.filter(wilaya=wilaya_filter)
    
    # Get distinct wilayas from orders
    wilayas = Order.objects.values_list('wilaya', flat=True).distinct().order_by('wilaya')
    
    return render(request, 'orders_list.html', {
        'orders': orders,
        'wilayas': wilayas
    })


@login_required
def update_order_status(request, order_id):
    """Update order status"""
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        new_status = request.POST.get('status')
        
        if new_status in ['pending', 'confirmed', 'cancelled', 'completed']:
            order.status = new_status
            order.save()
            
            status_names = {
                'pending': 'قيد الانتظار',
                'confirmed': 'مؤكد',
                'cancelled': 'ملغي',
                'completed': 'مكتمل'
            }
            messages.success(request, f'تم تحديث حالة الطلب إلى: {status_names[new_status]}')
        else:
            messages.error(request, 'حالة غير صحيحة')
    
    return redirect('orders_list')


@login_required
def export_orders(request):
    """Export selected orders to Excel for transporter"""
    if request.method == 'POST':
        order_ids = request.POST.getlist('order_ids')
        
        if not order_ids:
            messages.error(request, 'الرجاء اختيار طلب واحد على الأقل')
            return redirect('orders_list')
        
        # Get selected orders
        orders = Order.objects.filter(id__in=order_ids).order_by('wilaya', 'ville')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "قائمة الطلبات"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 15
        
        # Headers
        headers = ['#', 'الاسم', 'رقم الهاتف', 'الولاية', 'المدينة', 'الحالة', 'المنتجات', 'المجموع']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row = 2
        for order in orders:
            # Get order items as text
            items_text = "\n".join([
                f"{item.produit.nom} × {item.quantite}" 
                for item in order.orderitem_set.all()
            ])
            
            # Status translation
            status_dict = {
                'pending': 'قيد الانتظار',
                'confirmed': 'مؤكد',
                'cancelled': 'ملغي',
                'completed': 'مكتمل'
            }
            
            ws.cell(row=row, column=1, value=order.id).border = border
            ws.cell(row=row, column=2, value=order.nom).border = border
            ws.cell(row=row, column=3, value=order.telephone).border = border
            ws.cell(row=row, column=4, value=order.wilaya).border = border
            ws.cell(row=row, column=5, value=order.ville).border = border
            ws.cell(row=row, column=6, value=status_dict.get(order.status, order.status)).border = border
            
            items_cell = ws.cell(row=row, column=7, value=items_text)
            items_cell.border = border
            items_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row=row, column=8, value=f"{order.total()} د.ت").border = border
            
            row += 1
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="orders_export_{date.today()}.xlsx"'
        wb.save(response)
        
        return response
    
    return redirect('orders_list')
